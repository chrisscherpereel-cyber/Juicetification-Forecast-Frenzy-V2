"""
Juicetification: Forecast Frenzy
==============================
(The Campus Juice Bar forecasting lab)
A guided, ~60-minute EXPERIENTIAL forecasting lab for an intro Operations
Management course.

Behaviors
  * Every input states what to type and WHICH period is being forecast.
  * Unique scenario per session (random Scenario ID).
  * Worked data shown as full Excel-style grids (column letters + row numbers).
  * Excel formulas are EVALUATED for correctness AND must use cell references —
    hardcoding a value that exists in the grid is rejected.
  * Seasonality: students compute the day average and the overall (grand)
    average themselves before forming the index.
  * Module order: … Accuracy (MAD/MAPE) → Model Selection, so students learn to
    read the error before choosing a method — and they identify the lowest error
    themselves.
  * Run the Bar: students calculate the plan numbers, then implement.

Run with:  streamlit run forecast_frenzy.py
"""

import datetime as dt
import io
import json
import random
import re
import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

import student_store as store
from juice_director import resolve_config, serve_manifest_if_requested
from manifest import MANIFEST

# set_page_config must be the first Streamlit command executed on the page.
st.set_page_config(page_title="Juicetification: Forecast Frenzy", page_icon="🥤", layout="wide")
serve_manifest_if_requested(MANIFEST)   # ?manifest=1 emits the schema JSON and stops
CFG, CTX = resolve_config(MANIFEST)      # instructor overrides via ?cfg= or ?game=; else defaults

# Student identity (only active when student_store is configured; safe no-ops otherwise).
game = store.game_code()
sid = store.get_student_id()
if store.enabled() and sid is None:
    # Sign-in gate: no scenario is created until a student id exists.
    st.title("🥤 Juicetification: Forecast Frenzy")
    st.markdown("**Sign in to save and resume your progress.**")
    _entered = st.text_input("Enter your student ID to begin", key="_gate_sid",
                             placeholder="your campus ID")
    if st.button("Start ▶", type="primary") and _entered.strip():
        store.set_student_id(_entered)
        st.rerun()
    st.stop()

# ----------------------------------------------------------------------------
# Economics (per juice) — instructor-configurable through the Director manifest
# ----------------------------------------------------------------------------
PRICE = CFG["price"]
VAR_COST = CFG["var_cost"]
CONTRIB = PRICE - VAR_COST
FRUIT_PREP_COST = CFG["fruit_prep_cost"]
BOTTLE_COST = CFG["bottle_cost"]
WAGE_PER_HR = CFG["wage_per_hr"]
SHIFT_HOURS = CFG["shift_hours"]
SERVICE_PER_EMP = CFG["service_per_emp"]
EMP_DAILY_CAP = SERVICE_PER_EMP * SHIFT_HOURS   # customers/employee/day
PROMO_COST = CFG["promo_cost"]
PROMO_LIFT = CFG["promo_lift"]
MOBILE_PREMIUM = 0.50
SATISFACTION_PENALTY = CFG["satisfaction_penalty"]

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DOW_INDEX = {"Mon": 1.05, "Tue": 1.00, "Wed": 1.10, "Thu": 1.08,
             "Fri": 0.95, "Sat": 0.55, "Sun": 0.60}
BASE_DEMAND = CFG["base_demand"]
REG = {"intercept": 40, "temp": 2.5, "promo": 45, "attend": 1.8}   # given equation
BAR_NAME = "Juicetification"

# Every required step in the lab (used for the progress bar & completeness score).
MODULES = [
    ("Warm-up (Start Here)", ["s_pred"]),
    ("Module 1 — Forecasting", ["r1_guess", "r1_why"]),
    ("Module 2 — Qualitative", ["r2_fc", "r2_when"]),
    ("Module 3 — Naïve", ["r3_sat", "r3_err", "r3_use"]),
    ("Module 4 — Moving average", ["r4_ma3", "r4_ma3b", "r4_explore", "r4_tradeoff"]),
    ("Module 5 — Exp. smoothing", ["r5_es1", "r5_es2", "r5_explore", "r5_alpha"]),
    ("Module 6 — Seasonality", ["r6_wedavg", "r6_grand", "r6_idx", "r6_fc", "r6_satavg",
                                "r6_satidx", "r6_why"]),
    ("Module 7 — Regression", ["r7_pred", "r7_pred2", "r7_driver"]),
    ("Module 8 — Accuracy", ["r9_mad", "r9_mape", "r9_interpret"]),
    ("Module 9 — Model selection", ["msel_identify", "msel_pick", "msel_defend"]),
    (f"Run {BAR_NAME}", ["rb_result", "rb_lesson"]),
    ("Debrief", ["dbf_sowhat", "dbf_nowhat"]),
    ("Capstone", ["cap_expand"]),
]
REQUIRED = [q for _name, qs in MODULES for q in qs]


# ----------------------------------------------------------------------------
# Per-session UNIQUE teaching data (clean numbers, checkable answers)
# ----------------------------------------------------------------------------
def make_lab_data(seed, base_demand=None):
    base_demand = BASE_DEMAND if base_demand is None else base_demand
    rng = np.random.default_rng(seed + 7)
    r5 = lambda a, b: int(round(rng.uniform(a, b) / 5) * 5)
    r10 = lambda a, b: int(round(rng.uniform(a, b) / 10) * 10)
    week = {d: r5(280, 335) for d in ["Mon", "Tue", "Wed", "Thu", "Fri"]}
    sat_actual = r5(320, 360)
    es_alpha = float(rng.choice([0.2, 0.3, 0.4]))
    es_F = r5(300, 320)
    es_A = es_F + r10(20, 40)
    es_A2 = es_F - r10(10, 30)
    # Seasonality: 4 weeks of daily demand — students compute the averages.
    seas_grid = []
    for _wk in range(4):
        seas_grid.append([int(round((base_demand * DOW_INDEX[d] + rng.normal(0, 8)) / 5) * 5)
                          for d in DAYS])
    wed_i, sat_i = DAYS.index("Wed"), DAYS.index("Sat")
    wed_col = [seas_grid[w][wed_i] for w in range(4)]
    sat_col = [seas_grid[w][sat_i] for w in range(4)]
    all_vals = [v for row in seas_grid for v in row]
    wed_avg = sum(wed_col) / 4
    sat_avg = sum(sat_col) / 4
    grand = sum(all_vals) / len(all_vals)
    r1 = (r5(75, 90), 1, r5(80, 92))
    r2 = (r5(50, 62), 0, r5(55, 68))
    acts = [r10(180, 260), r10(230, 300), r10(280, 340), r10(360, 430)]
    errs = [int(x) for x in rng.choice([-30, -20, -10, 10, 20, 30], size=4, replace=True)]
    fcs = [int(a + e) for a, e in zip(acts, errs)]
    return {"week": week, "sat_actual": sat_actual,
            "es_alpha": es_alpha, "es_F": es_F, "es_A": es_A, "es_A2": es_A2,
            "seas_grid": seas_grid, "wed_avg": wed_avg, "sat_avg": sat_avg, "grand": grand,
            "base": base_demand, "r1": r1, "r2": r2, "acts": acts, "fcs": fcs}


@st.cache_data
def simulate_semester(seed=42, n_days=84, base_demand=BASE_DEMAND):
    rng = np.random.default_rng(seed)
    rows = []
    for d in range(n_days):
        dow = DAYS[d % 7]
        week = d // 7 + 1
        temp = float(np.clip(rng.normal(72 + 0.15 * d, 8), 45, 100))
        weather = 1.0 + 0.010 * (temp - 72)
        exam = 1 if week in (6, 12) else 0
        event = 1 if rng.random() < 0.12 else 0
        attendance = float(np.clip(rng.normal(82, 9) * DOW_INDEX[dow], 20, 100))
        promo = 1 if rng.random() < 0.18 else 0
        mean = (base_demand * DOW_INDEX[dow] * weather * (1.18 if exam else 1)
                * (1.35 if event else 1) * (attendance / 82.0) * (1.22 if promo else 1))
        demand = int(max(0, rng.normal(mean, mean * 0.08)))
        rows.append({"day": d + 1, "week": week, "dow": dow, "temp_f": round(temp, 1),
                     "exam": exam, "event": event, "attendance": round(attendance, 1),
                     "promo": promo, "demand": demand})
    return pd.DataFrame(rows)


def naive_forecast(s): return s.shift(1)
def moving_average(s, w): return s.shift(1).rolling(w).mean()


def exp_smoothing(s, alpha):
    v = s.values.astype(float); fc = np.full(len(v), np.nan)
    if len(v):
        level = v[0]
        for t in range(1, len(v)):
            fc[t] = level; level = alpha * v[t] + (1 - alpha) * level
    return pd.Series(fc, index=s.index)


def seasonal_indices(df):
    return (df.groupby("dow")["demand"].mean() / df["demand"].mean()).reindex(DAYS).to_dict()


def seasonal_naive_forecast(df):
    idx = seasonal_indices(df)
    return df["demand"].shift(1).rolling(7).mean() * df["dow"].map(idx)


def regression_fit(df):
    X = np.column_stack([np.ones(len(df)), df["temp_f"], df["promo"], df["attendance"],
                         df["exam"], df["event"]])
    y = df["demand"].values.astype(float)
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    return beta, X @ beta


def mad(a, f):
    a, f = np.asarray(a, float), np.asarray(f, float); m = ~np.isnan(f)
    return float(np.mean(np.abs(a[m] - f[m]))) if m.any() else np.nan


def mape(a, f):
    a, f = np.asarray(a, float), np.asarray(f, float); m = (~np.isnan(f)) & (a != 0)
    return float(np.mean(np.abs((a[m] - f[m]) / a[m])) * 100) if m.any() else np.nan


def rmse(a, f):
    a, f = np.asarray(a, float), np.asarray(f, float); m = ~np.isnan(f)
    return float(np.sqrt(np.mean((a[m] - f[m]) ** 2))) if m.any() else np.nan


def run_day_pnl(demand, employees, fruit_prep, bottles, promo, mobile_reserve):
    service_capacity = employees * EMP_DAILY_CAP
    realized = int(demand * (1 + PROMO_LIFT if promo else 1.0))
    servable = min(realized, service_capacity, bottles + fruit_prep)
    mobile_served = min(int(realized * 0.25), mobile_reserve, servable)
    units = servable
    revenue = units * PRICE + mobile_served * MOBILE_PREMIUM
    labor = employees * SHIFT_HOURS * WAGE_PER_HR
    bottles_sold = min(bottles, units)
    ing = units * VAR_COST
    fruit_waste = max(0, fruit_prep - (units - bottles_sold)) * FRUIT_PREP_COST
    bottle_waste = max(0, bottles - bottles_sold) * BOTTLE_COST
    promo_cost = PROMO_COST if promo else 0.0
    unmet = max(0, realized - units)
    goodwill = unmet * SATISFACTION_PENALTY
    profit = revenue - labor - ing - fruit_waste - bottle_waste - promo_cost - goodwill
    return {"realized_demand": realized, "units_sold": units, "unmet_demand": unmet,
            "revenue": round(revenue, 2), "labor_cost": round(labor, 2),
            "ingredient_cost": round(ing, 2), "fruit_waste": round(fruit_waste, 2),
            "bottle_waste": round(bottle_waste, 2), "promo_cost": round(promo_cost, 2),
            "goodwill_lost": round(goodwill, 2), "profit": round(profit, 2),
            "satisfaction": round(100 * units / realized if realized else 100, 1)}


# ----------------------------------------------------------------------------
# Excel formula evaluation + "must use cell references" check
# ----------------------------------------------------------------------------
def eval_excel(formula, cells):
    from openpyxl.utils import column_index_from_string, get_column_letter
    if not formula:
        return None
    f = formula.strip()
    if not f.startswith("="):
        return None
    f = f[1:].upper().replace(" ", "")
    if not f:
        return None
    try:
        def rng(m):
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            i1, i2 = column_index_from_string(c1), column_index_from_string(c2)
            vals = []
            for cc in range(min(i1, i2), max(i1, i2) + 1):
                for rr in range(min(r1, r2), max(r1, r2) + 1):
                    k = f"{get_column_letter(cc)}{rr}"
                    if k not in cells:
                        raise KeyError(k)
                    vals.append(float(cells[k]))
            return "ARR([" + ",".join(repr(v) for v in vals) + "])"

        f = re.sub(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", rng, f)

        def one(m):
            k = m.group(0)
            if k not in cells:
                raise KeyError(k)
            return repr(float(cells[k]))

        f = re.sub(r"[A-Z]+\d+", one, f)
        NS = {
            "ARR": lambda x: np.asarray(x, dtype=float),
            "ABS": np.abs,
            "AVERAGE": lambda *a: float(np.mean(np.concatenate(
                [np.atleast_1d(np.asarray(x, float)) for x in a]))),
            "SUM": lambda *a: float(np.sum([np.sum(np.asarray(x, float)) for x in a])),
            "SUMPRODUCT": lambda *a: (float(np.sum(np.asarray(a[0], float))) if len(a) == 1
                                      else float(np.sum(np.prod(
                                          [np.asarray(x, float) for x in a], axis=0)))),
            "COUNT": lambda *a: float(sum(np.atleast_1d(np.asarray(x, float)).size for x in a)),
            "MIN": lambda *a: float(np.min([np.min(np.asarray(x, float)) for x in a])),
            "MAX": lambda *a: float(np.max([np.max(np.asarray(x, float)) for x in a])),
            "ROUND": lambda x, n=0: float(round(float(x), int(n))),
        }
        val = eval(f, {"__builtins__": {}}, NS)
        arr = np.ravel(np.asarray(val, dtype=float))
        return float(arr[0]) if arr.size else None
    except Exception:
        return None


def hardcoded_data_value(formula, cells):
    """Return a numeric literal the student typed that duplicates a value available
    as a cell (so they should have used the cell reference). None if clean."""
    if not formula:
        return None
    f = formula.strip().upper().replace(" ", "")
    if not f.startswith("="):
        return None
    f = f[1:]
    f = re.sub(r"[A-Z]+\d+:[A-Z]+\d+", " ", f)   # strip ranges
    f = re.sub(r"[A-Z]+\d+", " ", f)             # strip single cell refs
    literals = re.findall(r"\d+\.?\d*", f)
    cellvals = {round(float(v), 4) for v in cells.values()}
    for lit in literals:
        try:
            x = round(float(lit), 4)
        except ValueError:
            continue
        if x in cellvals:
            return lit
    return None


# ============================================================================
# Companion Excel workbook (generated in-app, downloadable) — coords match app
# ============================================================================
@st.cache_data
def build_workbook(seed, base_demand=BASE_DEMAND):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    L = make_lab_data(seed, base_demand); W = L["week"]
    ARIAL = "Arial"
    tf = Font(name=ARIAL, size=14, bold=True); hf = Font(name=ARIAL, size=11, bold=True)
    nf = Font(name=ARIAL, size=11); note = Font(name=ARIAL, size=10, italic=True, color="555555")
    yel = PatternFill("solid", fgColor="FFF2CC"); blu = PatternFill("solid", fgColor="DDEBF7")
    grey = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="BBBBBB"); bx = Border(thin, thin, thin, thin)
    ctr = Alignment(horizontal="center")
    wb = Workbook()

    def setup(ws, t, sub):
        ws["A1"] = t; ws["A1"].font = tf; ws["A2"] = sub; ws["A2"].font = note
        ws.column_dimensions["A"].width = 40
        for c in "BCDEFGH": ws.column_dimensions[c].width = 12

    def data(ws, ref, v):
        ws[ref] = v; ws[ref].fill = blu; ws[ref].border = bx; ws[ref].font = nf; ws[ref].alignment = ctr

    def inp(ws, ref):
        ws[ref].fill = yel; ws[ref].border = bx; ws[ref].font = nf; ws[ref].alignment = ctr

    ws = wb.active; ws.title = "README"
    setup(ws, "Juicetification: Forecast Frenzy — Excel Practice Workbook",
          f"Scenario ID {seed}. Type a FORMULA (start with =) in every YELLOW cell. Use CELL "
          "REFERENCES, never typed-in numbers.")
    for i, t in enumerate([
        "", "LEGEND:  Yellow = you fill in (write a formula)   |   Blue = given data",
        "", "Naive:           = the previous day's cell",
        "Moving average:  =AVERAGE(range)",
        "Exp. smoothing:  =Fprev + alpha*(Actual - Fprev)     (reference the alpha cell)",
        "Seasonal index:  day average and overall average are AVERAGE() formulas you build,",
        "                 then index = dayAvgCell / overallAvgCell ; Forecast = baseCell * indexCell",
        "Regression:      =interceptCell + tempCoefCell*Temp + promoCoefCell*Promo + attCoefCell*Att",
        "MAD:             =AVERAGE(|error| column)    MAPE = AVERAGE(|error|/actual)*100",
        "", "Cell references here MATCH the app. Check yourself on the 'Answer Key' tab."],
            start=4):
        ws[f"A{i}"] = t; ws[f"A{i}"].font = nf
    ws.column_dimensions["A"].width = 96

    for name, sub in [("Naive", "F(t+1)=A(t)"), ("MovingAvg", "F(t+1)=average of last n days")]:
        ws = wb.create_sheet(name); setup(ws, name, sub)
        dd = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        vv = [W["Mon"], W["Tue"], W["Wed"], W["Thu"], W["Fri"], L["sat_actual"]]
        ws["A4"] = "Day"; ws["A4"].font = hf; ws["A5"] = "Actual"; ws["A5"].font = hf
        for j, (d, v) in enumerate(zip(dd, vv)):
            c = get_column_letter(2 + j)
            ws[f"{c}4"] = d; ws[f"{c}4"].font = hf; ws[f"{c}4"].alignment = ctr
            data(ws, f"{c}5", v)
        if name == "Naive":
            ws["A7"] = "Saturday naive forecast (=Fri)"; ws["A7"].font = nf; inp(ws, "B7")
            ws["A8"] = "Thursday abs error =ABS(Thu-Wed)"; ws["A8"].font = nf; inp(ws, "B8")
        else:
            ws["A7"] = "3-day MA forecast for Saturday"; ws["A7"].font = nf; inp(ws, "B7")
            ws["A8"] = "3-day MA forecast for Sunday"; ws["A8"].font = nf; inp(ws, "B8")

    ws = wb.create_sheet("ExpSmoothing"); setup(ws, "Exp. smoothing", "F(t+1)=F(t)+alpha*(A(t)-F(t))")
    ws["A4"] = "alpha"; ws["A4"].font = hf; data(ws, "B4", L["es_alpha"])
    ws["A6"] = "Day"; ws["B6"] = "Prior F"; ws["C6"] = "Actual"; ws["D6"] = "New forecast (next day)"
    for c in "ABCD": ws[f"{c}6"].font = hf
    ws["A7"] = "Day 1"; ws["A7"].font = nf; data(ws, "B7", L["es_F"]); data(ws, "C7", L["es_A"]); inp(ws, "D7")
    ws["A8"] = "Day 2"; ws["A8"].font = nf; data(ws, "C8", L["es_A2"]); inp(ws, "D8")
    ws["A10"] = "Day-2 prior forecast is your Day-1 answer (D7). Hint: =B7+$B$4*(C7-B7)"; ws["A10"].font = note

    ws = wb.create_sheet("Seasonality"); setup(ws, "Seasonality", "You compute the averages first.")
    ws["A4"] = "Week"; ws["A4"].font = hf
    for j, d in enumerate(DAYS):
        c = get_column_letter(2 + j); ws[f"{c}4"] = d; ws[f"{c}4"].font = hf; ws[f"{c}4"].alignment = ctr
    for wk in range(4):
        ws[f"A{5+wk}"] = f"Wk {wk+1}"; ws[f"A{5+wk}"].font = nf
        for j, v in enumerate(L["seas_grid"][wk]):
            data(ws, f"{get_column_letter(2+j)}{5+wk}", v)
    labels = [("Wed average  =AVERAGE(D5:D8)", "B10"), ("Sat average  =AVERAGE(G5:G8)", "B11"),
              ("Overall (grand) average  =AVERAGE(B5:H8)", "B12"), ("Base level", "B13"),
              ("Wed index  =B10/B12", "B14"), ("Sat index  =B11/B12", "B15"),
              ("Wed forecast  =B13*B14", "B16")]
    for lab, ref in labels:
        r = int(ref[1:]); ws[f"A{r}"] = lab; ws[f"A{r}"].font = nf
        if ref == "B13":
            data(ws, ref, L["base"])
        else:
            inp(ws, ref)

    ws = wb.create_sheet("Regression"); setup(ws, "Regression", "Demand=40+2.5*Temp+45*Promo+1.8*Att")
    for i, (lab, v) in enumerate([("Intercept", REG["intercept"]), ("Temp coef", REG["temp"]),
                                  ("Promo coef", REG["promo"]), ("Attend coef", REG["attend"])], start=4):
        ws[f"A{i}"] = lab; ws[f"A{i}"].font = nf; data(ws, f"B{i}", v)
    ws["A9"] = "Temp"; ws["B9"] = "Promo"; ws["C9"] = "Attendance"; ws["D9"] = "Predicted"
    for c in "ABCD": ws[f"{c}9"].font = hf
    for i, row in enumerate([L["r1"], L["r2"]], start=10):
        for j, v in enumerate(row): data(ws, f"{get_column_letter(1+j)}{i}", v)
        inp(ws, f"D{i}")
    ws["A13"] = "Hint: =$B$4+$B$5*A10+$B$6*B10+$B$7*C10 (reference the coefficient cells)"
    ws["A13"].font = note

    ws = wb.create_sheet("Accuracy"); setup(ws, "Accuracy", "MAD=avg|err| ; MAPE=avg(|err|/act)*100")
    for j, h in enumerate(["Day", "Actual", "Forecast", "|Error|", "%Error"]):
        c = get_column_letter(1 + j); ws[f"{c}4"] = h; ws[f"{c}4"].font = hf; ws[f"{c}4"].alignment = ctr
    for i in range(4):
        r = 5 + i
        ws[f"A{r}"] = i + 1; ws[f"A{r}"].font = nf; ws[f"A{r}"].alignment = ctr
        data(ws, f"B{r}", L["acts"][i]); data(ws, f"C{r}", L["fcs"][i]); inp(ws, f"D{r}"); inp(ws, f"E{r}")
    ws["A10"] = "MAD (customers)"; ws["A10"].font = hf; inp(ws, "B10")
    ws["A11"] = "MAPE (%)"; ws["A11"].font = hf; inp(ws, "B11")
    ws["A13"] = "Hints: D5 =ABS(B5-C5) ; E5 =D5/B5 ; MAD =AVERAGE(D5:D8) ; MAPE =AVERAGE(E5:E8)*100"
    ws["A13"].font = note

    ws = wb.create_sheet("Answer Key"); setup(ws, "Answer Key", "Formulas reference each tab's cells.")
    ws["A4"] = "Item"; ws["B4"] = "Model formula"; ws["C4"] = "Result"
    for c in "ABC": ws[f"{c}4"].font = hf; ws[f"{c}4"].fill = grey
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 32; ws.column_dimensions["C"].width = 12
    key = [
        ("Naive Sat", " =F5", "=Naive!F5"),
        ("Naive Thu abs error", " =ABS(E5-D5)", "=ABS(Naive!E5-Naive!D5)"),
        ("MA3 Saturday", " =AVERAGE(D5:F5)", "=AVERAGE(MovingAvg!D5:F5)"),
        ("MA3 Sunday", " =AVERAGE(E5:G5)", "=AVERAGE(MovingAvg!E5:G5)"),
        ("Exp smoothing Day1", " =B7+$B$4*(C7-B7)",
         "=ExpSmoothing!B7+ExpSmoothing!B4*(ExpSmoothing!C7-ExpSmoothing!B7)"),
        ("Seas Wed average", " =AVERAGE(D5:D8)", "=AVERAGE(Seasonality!D5:D8)"),
        ("Seas overall average", " =AVERAGE(B5:H8)", "=AVERAGE(Seasonality!B5:H8)"),
        ("Seas Wed index", " =B10/B12", "=AVERAGE(Seasonality!D5:D8)/AVERAGE(Seasonality!B5:H8)"),
        ("Seas Wed forecast", " =B13*B14",
         "=Seasonality!B13*(AVERAGE(Seasonality!D5:D8)/AVERAGE(Seasonality!B5:H8))"),
        ("Regression row1", " =B4+B5*A10+B6*B10+B7*C10",
         "=Regression!B4+Regression!B5*Regression!A10+Regression!B6*Regression!B10+Regression!B7*Regression!C10"),
        ("MAD", " =AVERAGE(D5:D8)",
         "=AVERAGE(ABS(Accuracy!B5-Accuracy!C5),ABS(Accuracy!B6-Accuracy!C6),ABS(Accuracy!B7-Accuracy!C7),ABS(Accuracy!B8-Accuracy!C8))"),
        ("MAPE", " =AVERAGE(E5:E8)*100",
         "=AVERAGE(ABS(Accuracy!B5-Accuracy!C5)/Accuracy!B5,ABS(Accuracy!B6-Accuracy!C6)/Accuracy!B6,ABS(Accuracy!B7-Accuracy!C7)/Accuracy!B7,ABS(Accuracy!B8-Accuracy!C8)/Accuracy!B8)*100"),
    ]
    for i, (lab, ftxt, fcell) in enumerate(key, start=5):
        ws[f"A{i}"] = lab; ws[f"A{i}"].font = nf
        ws[f"B{i}"] = ftxt; ws[f"B{i}"].font = Font(name="Consolas", size=10)
        ws[f"C{i}"] = fcell; ws[f"C{i}"].font = nf; ws[f"C{i}"].alignment = ctr
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ============================================================================
# Submittable PDF summary (generated in-app)
# ============================================================================
_PDF_REPL = {"’": "'", "‘": "'", "“": '"', "”": '"', "–": "-", "—": "-", "…": "...",
             "•": "-", "≈": "~", "×": "x", "°": "deg", "α": "alpha", "✅": "[correct]",
             "❌": "[incorrect]", "⬜": "[ ]", "🎯": "", "🔁": "", "📊": "", "➡️": "", "💡": "",
             "🥤": "", "🏪": "", "📎": "", "🎲": "", "🔎": "", "✍️": "", "📋": "", "✏️": "",
             "🤔": "", "👁️": "", "▶": ">"}


def _ascii(s):
    s = str(s)
    for k, v in _PDF_REPL.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "ignore").decode("latin-1")


def _pdf_escape(s):
    return s.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def build_pdf(student, seed, method, order, responses, mastery_line=None, perf_line=None):
    """Pure-Python PDF writer (no third-party dependency) — a simple multi-page
    text report using the standard Helvetica fonts."""
    lines = []

    def add(text, size=10, bold=False):
        lines.append((text, size, bold))

    checks = [v for v in responses.values() if v["correct"] is not None]
    n_ok = sum(1 for v in checks if v["correct"])
    done_steps = sum(1 for q in REQUIRED if q in responses)
    score = int(round(100 * done_steps / len(REQUIRED)))
    add("Juicetification: Forecast Frenzy", 16, True)
    add("Campus juice bar - forecasting lab", 10, False)
    add("", 4)
    add(f"Student: {student}", 11)
    add(f"Session ID: {seed}", 11)
    add(f"Date: {dt.date.today().isoformat()}", 11)
    add(f"Method selected: {method}", 11)
    add(f"Completeness score: {score}/100  ({done_steps}/{len(REQUIRED)} steps)", 11)
    add(mastery_line or "Mastery score: not attempted", 11)
    add(perf_line or "Performance score: not attempted", 11)
    add(f"Checked calculations: {n_ok}/{len(checks)} correct" if checks
        else "Checked calculations: none attempted", 11)
    add("", 4)
    cur = None
    for qid in order:
        v = responses[qid]
        if v["section"] != cur:
            cur = v["section"]; add("", 3); add(cur, 12, True)
        mark = "" if v["correct"] is None else (" [correct]" if v["correct"] else " [incorrect]")
        add(f"- {v['label']}: {v['answer']}{mark}", 10)

    def wrap(text, size):
        if not text:
            return [""]
        maxchars = max(12, int(512 / (size * 0.52)))
        out, line = [], ""
        for word in text.split(" "):
            while len(word) > maxchars:              # hard-break very long tokens
                if line:
                    out.append(line); line = ""
                out.append(word[:maxchars]); word = word[maxchars:]
            if len(line) + len(word) + 1 <= maxchars:
                line = (line + " " + word).strip()
            else:
                if line:
                    out.append(line)
                line = word
        out.append(line)
        return out

    # Lay text out into pages of drawing operators.
    pages, ops, y = [], [], 750.0
    for text, size, bold in lines:
        for seg in wrap(text, size):
            lh = size * 1.35 if seg else size * 0.9
            if y - lh < 50:
                pages.append(ops); ops = []; y = 750.0
            if seg:
                font = "F2" if bold else "F1"
                ops.append(f"BT /{font} {size} Tf 50 {y:.1f} Td ({_pdf_escape(_ascii(seg))}) Tj ET")
            y -= lh
    pages.append(ops)

    # Assemble PDF objects.
    n = len(pages)
    page_ids = [5 + 2 * i for i in range(n)]
    content_ids = [6 + 2 * i for i in range(n)]
    objects = {
        1: "<< /Type /Catalog /Pages 2 0 R >>",
        2: f"<< /Type /Pages /Kids [{' '.join(f'{p} 0 R' for p in page_ids)}] /Count {n} >>",
        3: "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        4: "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
    }
    for i in range(n):
        stream = "\n".join(pages[i])
        objects[content_ids[i]] = f"<< /Length {len(stream)} >>\nstream\n{stream}\nendstream"
        objects[page_ids[i]] = ("<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
                                "/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> "
                                f"/Contents {content_ids[i]} 0 R >>")
    out = b"%PDF-1.4\n"
    offsets = {}
    for num in sorted(objects):
        offsets[num] = len(out)
        out += f"{num} 0 obj\n{objects[num]}\nendobj\n".encode("latin-1")
    xref_pos = len(out)
    maxid = max(objects)
    out += f"xref\n0 {maxid + 1}\n".encode("latin-1")
    out += b"0000000000 65535 f \n"
    for num in range(1, maxid + 1):
        out += (f"{offsets[num]:010d} 00000 n \n" if num in offsets
                else "0000000000 65535 f \n").encode("latin-1")
    out += (f"trailer\n<< /Size {maxid + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n"
            "%%EOF").encode("latin-1")
    return out


# ============================================================================
# Response capture + reusable teaching widgets
# ============================================================================
def _init_state():
    st.session_state.setdefault("responses", {})
    st.session_state.setdefault("order", [])
    st.session_state.setdefault("student", "")
    st.session_state.setdefault("section", "")
    st.session_state.setdefault("first_try", {})   # qid -> bool (first attempt correct?)
    st.session_state.setdefault("runs", [])        # Run-Juicetification day results
    if "seed" not in st.session_state:
        if CTX["seed"] is not None:                       # instructor fixed a class-wide seed
            st.session_state["seed"] = CTX["seed"]
        elif sid is not None:                             # stable, unique scenario per student
            st.session_state["seed"] = store.derive_seed(game, sid, lo=1000, hi=9999)
        else:                                             # unconfigured: original random Session ID
            st.session_state["seed"] = random.randint(1000, 9999)


def record_first(key, ok):
    """Remember whether the FIRST attempt at a checkable item was correct."""
    ft = st.session_state.setdefault("first_try", {})
    if key not in ft:
        ft[key] = bool(ok)


def mastery_score():
    ft = st.session_state.get("first_try", {})
    if not ft:
        return None, 0, 0
    correct = sum(1 for v in ft.values() if v)
    return int(round(100 * correct / len(ft))), correct, len(ft)


def performance_score():
    runs = st.session_state.get("runs", [])
    if not runs:
        return None, runs
    return int(round(sum(r["perf"] for r in runs) / len(runs))), runs


# Session_state keys that carry student progress (persisted / restored via student_store).
PROGRESS_KEYS = ["responses", "order", "first_try", "runs", "chosen_method", "plan_fc", "student"]


def _jsonable(v):
    """Coerce a value into something json.dumps can handle (numpy, sets, DataFrames)."""
    if isinstance(v, dict):
        return {k: _jsonable(x) for k, x in v.items()}
    if isinstance(v, (list, tuple)):
        return [_jsonable(x) for x in v]
    if isinstance(v, set):
        return sorted(_jsonable(x) for x in v)
    if isinstance(v, np.ndarray):
        return v.tolist()
    if isinstance(v, np.generic):
        return v.item()
    if isinstance(v, pd.DataFrame):
        return v.to_dict("list")
    return v


def autosave():
    """Persist a plain-JSON snapshot of just the progress keys (no-op unless enabled).

    Debounced: reflect() calls save() on every rerun while text is present, so we skip
    the network write unless the snapshot actually changed since the last upload.
    """
    if not (store.enabled() and sid):
        return
    snap = {k: _jsonable(st.session_state[k]) for k in PROGRESS_KEYS if k in st.session_state}
    try:
        blob = json.dumps(snap, sort_keys=True, separators=(",", ":"))
    except Exception:
        return
    if st.session_state.get("_autosave_blob") == blob:
        return                                  # unchanged since last write; skip the upload
    try:
        store.save(game, sid, snap)
        st.session_state["_autosave_blob"] = blob
    except Exception:
        pass


def save(qid, label, answer, correct=None):
    if qid not in st.session_state["responses"]:
        st.session_state["order"].append(qid)
    st.session_state["responses"][qid] = {"section": st.session_state["section"],
                                           "label": label, "answer": answer, "correct": correct}
    autosave()


def answered(qid):
    return qid in st.session_state["responses"]


def excel_grid(col_letters, rows, start_row=1, label_cols=()):
    html = ['<table class="xl"><tr><th class="corner"></th>']
    html += [f'<th>{c}</th>' for c in col_letters] + ['</tr>']
    for i, row in enumerate(rows):
        rn = start_row + i
        html.append(f'<tr><td class="rownum">{rn}</td>')
        for ci, val in enumerate(row):
            cls = "lbl" if ci in label_cols else ""
            html.append(f'<td class="{cls}">{"" if val is None else val}</td>')
        html.append('</tr>')
    html.append('</table>')
    st.markdown("".join(html), unsafe_allow_html=True)


def num_task(qid, label, correct, worked_md, feedback_md, excel_model, excel_hint,
             cells, tol=0.03, units=""):
    ustr = f" {units}" if units else ""
    st.caption("Choose how to work this, then type your answer in the box.")
    ap = st.radio("approach", ["✏️ Compute it myself", "🤔 Estimate, then reveal",
                               "👁️ Show answer & interpret"], key=f"ap_{qid}",
                  horizontal=True, label_visibility="collapsed")
    if ap.startswith("✏️"):
        val = st.number_input(f"➡️ Enter {label}{(' (in '+units+')') if units else ''}:",
                              value=None, key=f"in_{qid}", placeholder="Type your number here")
        if st.button("Check my answer", key=f"btn_{qid}"):
            if val is None:
                st.warning("Type a number in the box above first.")
            else:
                ok = abs(val - correct) <= max(abs(correct) * tol, 0.6)
                (st.success if ok else st.error)(
                    f"{'✅ Correct' if ok else '❌ Not quite'} — the answer is {correct:g}{ustr}."
                    + ("" if ok else f" You entered {val:g}."))
                with st.expander("Worked solution", expanded=not ok):
                    st.markdown(worked_md)
                record_first(qid, ok)
                save(qid, label, f"{val:g}{ustr}", ok)
    elif ap.startswith("🤔"):
        val = st.number_input(f"➡️ Enter your estimate for {label}:", value=None,
                              key=f"in_{qid}", placeholder="Type your estimate")
        if st.button("Reveal worked solution", key=f"btn_{qid}"):
            st.info(f"Answer: **{correct:g}{ustr}**"); st.markdown(worked_md)
            save(qid, label, f"estimate {val:g}" if val is not None else "estimate (blank)", None)
    else:
        st.info(f"Answer: **{correct:g}{ustr}**")
        with st.expander("How it's calculated"):
            st.markdown(worked_md)
        if st.button("Got it — mark this done", key=f"btn_{qid}"):
            save(qid, label, "interpreted", None)

    st.markdown(f"**📊 Now in Excel** — type the formula using **cell references** (no typed-in "
                f"numbers). I'll run it on the grid. _Hint: {excel_hint}_")
    x1, x2 = st.columns([3, 1])
    xf = x1.text_input("excel", key=f"xl_{qid}", placeholder="=...", label_visibility="collapsed")
    if x2.button("Check formula", key=f"xlb_{qid}"):
        xff = xf.strip()
        hard = hardcoded_data_value(xff, cells)
        res = eval_excel(xff, cells)
        if not xff.startswith("="):
            st.error("❌ Start the formula with '='."); ok = False
        elif hard is not None:
            st.error(f"❌ Don't type the value **{hard}** — it's data shown in the grid. Reference "
                     "its **cell** instead (Excel should pull from cells, not hardcoded numbers).")
            ok = False
        elif res is None:
            st.error("❌ I couldn't evaluate that. Use the cell references shown and check parentheses.")
            ok = False
        elif abs(res - correct) <= max(abs(correct) * tol, 0.6):
            st.success(f"✅ Correct — your formula computes **{res:g}** from cell references.")
            ok = True
        else:
            st.error(f"❌ Your formula computes **{res:g}**, but the answer is **{correct:g}{ustr}**. "
                     "Check which cells/range you referenced.")
            ok = False
        st.markdown(f"**Model formula:** `{excel_model}`")
        record_first(f"{qid}_xl", ok)
        save(f"{qid}_xl", f"{label} — Excel formula", f"{xf} → {res}", ok)

    if answered(qid) and feedback_md:
        with st.expander("💡 Feedback — is your answer reasonable, and what's even better?",
                         expanded=True):
            st.markdown(feedback_md)


def reflect(qid, prompt, short_label, feedback_md=None, height=90, rubric=None):
    rubric = rubric or ["Names a specific cause or reason (not just 'it changes')",
                        "Names a trade-off, downside, or 'it depends'",
                        "Connects it to a decision, number, or action"]
    st.markdown(f"**✍️ Write your answer:** {prompt}")
    txt = st.text_area(prompt, key=f"rf_{qid}", height=height,
                       placeholder="Type your response here…", label_visibility="collapsed")
    if txt.strip():
        st.markdown("**Self-check — tick what your answer includes:**")
        met = sum(1 for i, crit in enumerate(rubric) if st.checkbox(crit, key=f"rub_{qid}_{i}"))
        save(qid, short_label, f"{txt.strip()}  [self-rubric {met}/{len(rubric)}]")
        if met == len(rubric):
            st.success("Complete answer — all rubric points covered. ✅")
        else:
            st.caption(f"{met}/{len(rubric)} rubric points ticked — use the strong-answer note below "
                       "to fill any gaps.")
        if feedback_md:
            with st.expander("💡 Feedback — compare your thinking to a strong answer", expanded=True):
                st.markdown(feedback_md)


def plan_calc(key, label, correct, formula_str, tol=0.05):
    """Lightweight numeric check for the Run-the-Bar plan (no grid)."""
    c1, c2 = st.columns([3, 1])
    v = c1.number_input(f"{label} — {formula_str}", value=None, key=f"pc_{key}",
                        placeholder="Type the result")
    if c2.button("Check", key=f"pcb_{key}"):
        if v is None:
            st.warning("Enter a number first.")
        else:
            ok = abs(v - correct) <= max(abs(correct) * tol, 0.5)
            (st.success if ok else st.error)(
                f"{'✅' if ok else '❌'} {label} = {correct:g}." + ("" if ok else f" You entered {v:g}."))
            record_first(f"plan_{key}", ok)
            save(f"plan_{key}", f"Plan — {label}", f"{v:g}", ok)


def overall_progress():
    resp = st.session_state["responses"]
    done = sum(1 for q in REQUIRED if q in resp)
    return done, len(REQUIRED)


def completion(required, next_label):
    have = [q for q in required if q in st.session_state["responses"]]
    st.divider()
    if len(have) == len(required):
        st.success(f"✅ **Section complete!** You've finished every item here — go to **{next_label}**.")
    else:
        st.info(f"⬜ **{len(have)} of {len(required)} done.** Finish the remaining "
                f"{len(required)-len(have)} item(s), then move to **{next_label}**.")


# ============================================================================
# UI  (set_page_config already called at the top, before the Director loader)
# ============================================================================
st.markdown("""<style>
table.xl{border-collapse:collapse;font-family:Calibri,Arial,sans-serif;font-size:13px;margin:6px 0}
table.xl td,table.xl th{border:1px solid #c6c6c6;padding:3px 10px;text-align:center;min-width:48px}
table.xl th{background:#eef1f5;color:#555;font-weight:600}
table.xl th.corner{background:#dfe3e8;min-width:26px}
table.xl td.rownum{background:#eef1f5;color:#555;font-weight:600;min-width:26px}
table.xl td.lbl{text-align:left;background:#fafafa;font-weight:600}
</style>""", unsafe_allow_html=True)
_init_state()

# Restore saved progress once per session (no-op unless enabled and identified).
if store.enabled() and sid and not st.session_state.get("_restored"):
    _saved = store.load(game, sid)
    if _saved:
        for _k in PROGRESS_KEYS:
            if _k in _saved:
                st.session_state[_k] = _saved[_k]
    st.session_state["_restored"] = True

seed = st.session_state["seed"]
L = make_lab_data(seed, BASE_DEMAND); W = L["week"]
df = simulate_semester(seed, 84, BASE_DEMAND)
train = df.iloc[:-14].copy(); holdout = 14

st.title("🥤 Juicetification: Forecast Frenzy")
st.markdown(f"**Manage {BAR_NAME}, the campus juice bar — a hands-on forecasting lab · ~60 min**")
if store.enabled() and sid:
    st.caption(f"Signed in as {sid} · progress saved automatically")

with st.sidebar:
    st.header("Your lab")
    st.session_state["student"] = st.text_input("Name / student ID", st.session_state["student"],
                                                 placeholder="Type your name")
    st.caption(f"Your workbook & scenario are unique to you  (#{seed}).")
    _done, _tot = overall_progress()
    _pct = int(round(100 * _done / _tot))
    st.progress(_done / _tot, text=f"Progress: {_pct}%  ({_done}/{_tot} steps)")
    _ms, _mc, _mt = mastery_score()
    if _ms is not None:
        st.caption(f"Mastery so far: {_ms}/100  ({_mc}/{_mt} right on first try)")
    if st.button("🔄 Start over (new scenario)"):
        for k in ["responses", "order", "seed", "first_try", "runs",
                  "_completion_recorded", "_completion_code", "_autosave_blob"]:
            st.session_state.pop(k, None)
        st.rerun()
    st.divider()
    st.download_button("⬇️ Excel practice workbook", build_workbook(seed, BASE_DEMAND),
                       file_name=f"Juicetification_Forecast_Frenzy_Practice_{seed}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       help="Same tasks & cell references, in a real spreadsheet.")
    st.divider()
    st.markdown("**Course objectives** — you'll be able to:")
    st.markdown("1. Explain forecasting & perishable service demand.\n2. Use qualitative signals.\n"
                "3. Compute naïve, moving-average & smoothing forecasts.\n4. Build & apply seasonal "
                "indices.\n5. Forecast with regression.\n6. Measure MAD & MAPE.\n7. Select & defend a "
                "method.\n8. Write Excel formulas with cell references.\n9. Turn a forecast into an "
                "operating plan.")

tabs = st.tabs(["📖 Start Here", "1 · Forecasting", "2 · Qualitative", "3 · Naïve", "4 · Moving Avg",
                "5 · Exp. Smoothing", "6 · Seasonality", "7 · Regression", "8 · Accuracy",
                "9 · Model Selection", f"🏪 Run {BAR_NAME}", "🎓 Debrief", "📝 Final Report"])

# Scroll the page to the top whenever the user switches tabs (screens). A unique nonce
# each rerun makes the injected HTML unique, so Streamlit remounts the iframe and re-binds
# to any freshly rendered tab buttons; binding is idempotent (data-stt flag) so handlers do
# not stack, and it scrolls ONLY on a tab click, never on ordinary edits/reruns.
components.html(
    "<script>(function(){var d=window.parent.document;"
    "function top(){try{window.parent.scrollTo(0,0);}catch(e){}"
    "var c=d.querySelector('section.main')||d.querySelector('[data-testid=\"stAppViewContainer\"]')||d.scrollingElement;"
    "if(c){c.scrollTop=0;}}"
    "d.querySelectorAll('button[role=\"tab\"]').forEach(function(b){"
    "if(!b.dataset.stt){b.dataset.stt='1';b.addEventListener('click',function(){setTimeout(top,40);});}});"
    "})();/*" + str(random.random()) + "*/</script>",
    height=0)


def objective_box(mins, text):
    st.info(f"🎯 **Learning objective** ({mins} min): {text}")


# ---- Start Here ----
with tabs[0]:
    st.session_state["section"] = "Start"
    st.subheader(f"Welcome, manager of {BAR_NAME}")
    st.markdown(
        f"You run **{BAR_NAME}**, the juice bar in the student union. Each morning you decide "
        "staffing, fruit prep, bottles, promos, and mobile capacity **before** you know how many "
        "customers show up. Get it right and you profit; get it wrong and you either dump spoiled "
        "fruit and pay idle staff, or sell out and lose customers.")
    st.markdown("> **Perishable service capacity:** an unmade juice at 2 p.m. can't be stored and "
                "sold tomorrow — that's why forecasting matters in services.")
    st.markdown("**How to work through this lab:** move through the tabs left → right. Each module "
                "follows the same rhythm — objective → formula → *you apply it* → *write the Excel "
                "formula with cell references* → **feedback** → a guiding question → then apply it "
                "**again**. When a tab is finished you'll see a green ✅ banner pointing you to the "
                "next one. Your progress bar is in the sidebar; your answers save automatically.")
    st.markdown("📎 **Your Excel practice workbook** is in the sidebar — its cell references match the "
                "grids in this app.")
    st.markdown(f"**Your demand history this term (Scenario {seed}):**")
    st.line_chart(df.set_index("day")["demand"])
    c1, c2, c3 = st.columns(3)
    c1.metric("Avg daily demand", f"{df['demand'].mean():.0f}")
    c2.metric("Busiest / quietest", f"{df['demand'].max()} / {df['demand'].min()}")
    c3.metric("Weekend vs weekday", "≈ 55–60%")
    reflect("s_pred",
            "Look at the chart above. In 2–3 sentences, describe the patterns you notice (the weekly "
            "rhythm, any spikes) and list what you think DRIVES the ups and downs.",
            "Warm-up — patterns & drivers",
            "A reasonable answer notices the weekly saw-tooth (weekends far lower) and a few spikes. A "
            "**stronger** answer names *causes* — day-of-week, temperature, exam weeks, campus events, "
            "promotions — and separates the weekly rhythm (seasonality) from one-off jumps (events).")
    completion(["s_pred"], "Tab 1 · Forecasting")

# ---- 1 Forecasting ----
with tabs[1]:
    st.session_state["section"] = "Module 1 — What forecasting is"
    st.subheader("Module 1 — What is forecasting?")
    objective_box(4, "Define forecasting and make a reasoned intuitive estimate.")
    st.markdown("**Forecasting** = estimating a future value from information you have now.")
    last = df.iloc[-1]
    st.markdown(f"**Your situation (Scenario {seed}).** Yesterday was **{last['dow']}**, "
                f"{last['temp_f']}°F, {'an EXAM week' if last['exam'] else 'a normal week'}, "
                f"{'a campus event was on' if last['event'] else 'no campus event'}. "
                f"Demand was **{last['demand']}** juices.")
    st.markdown("**➡️ Type your forecast for TOMORROW's demand (number of customers):**")
    g = st.number_input("Your forecast for tomorrow (customers)", value=None, key="r1g",
                        placeholder="Enter a whole number, e.g. 300", label_visibility="collapsed")
    if g is not None:
        save("r1_guess", "Intuitive forecast (tomorrow)", int(g))
        lo, hi = last["demand"] * 0.85, last["demand"] * 1.15
        if g < lo:
            fb = (f"**{int(g)}** is well below yesterday's {last['demand']} — reasonable only if you "
                  "expect a weekend/break. Otherwise it looks low.")
        elif g > hi:
            fb = (f"**{int(g)}** is well above yesterday's {last['demand']} — reasonable only if you "
                  "expect an event, promo, or hot exam day.")
        else:
            fb = (f"**{int(g)}** is within ±15% of yesterday's {last['demand']} — a sensible anchor.")
        with st.expander("💡 Feedback on your estimate", expanded=True):
            st.markdown(fb + "\n\n'Anchor on the last value, then adjust' is exactly the **naïve** "
                        "method you'll formalize in Module 3.")
    reflect("r1_why",
            "What specific information did you use to pick that number, and what would make you "
            "revise it up or down?", "Reasoning behind the estimate",
            "Strong answers cite *specific, observable* signals (it's a Wednesday, it's 88°F, it's "
            "finals) rather than gut feel, and say which signal would move the number most.")
    completion(["r1_guess", "r1_why"], "Tab 2 · Qualitative")

# ---- 2 Qualitative ----
with tabs[2]:
    st.session_state["section"] = "Module 2 — Qualitative forecasting"
    st.subheader("Module 2 — Qualitative forecasting")
    objective_box(5, "Turn given judgmental signals into a single forecast number.")
    st.markdown("With little/no data you lean on **judgment**: manager opinion, customer/staff input, "
                "and known events. Below is a briefing — you *forecast from it*.")
    base = df["demand"].tail(7).mean()
    rng2 = np.random.default_rng(seed + 3)
    mgr = str(rng2.choice(["Slow", "Normal", "Busy", "Slammed"]))
    event_line = str(rng2.choice(["a home football game", "a career fair in the union",
                                  "no special events", "a spring concert on the quad"]))
    comment = str(rng2.choice(["“the lines were too long last week”", "“love the new mango blend”",
                               "“we'll be back for finals”"]))
    st.info(f"**📋 Manager's briefing (Scenario {seed}):**\n\n"
            f"- Recent 7-day average demand: **{base:.0f}** juices/day\n"
            f"- Manager's read on next week: **“{mgr}”**\n"
            f"- Known events: **{event_line}**\n"
            f"- A common customer comment: {comment}")
    st.markdown("**➡️ Based only on this briefing, type your qualitative forecast for a TYPICAL DAY "
                "next week (customers):**")
    q = st.number_input("Your qualitative forecast (customers)", value=None, key="r2q",
                        placeholder="Enter a number", label_visibility="collapsed")
    mult = {"Slow": 0.87, "Normal": 1.0, "Busy": 1.15, "Slammed": 1.30}[mgr]
    reasonable = base * mult
    if q is not None:
        save("r2_fc", "Qualitative forecast (typical day next week)", int(q))
        lo, hi = reasonable * 0.9, reasonable * 1.12
        if lo <= q <= hi:
            verdict = (f"✅ **Reasonable.** A '{mgr}' read scales the {base:.0f} base to roughly "
                       f"**{reasonable:.0f}**, and you're in that zone.")
        elif q < lo:
            verdict = (f"⚠️ **A bit low.** A '{mgr}' read implies ≈ {reasonable:.0f}.")
        else:
            verdict = (f"⚠️ **A bit high.** A '{mgr}' read implies ≈ {reasonable:.0f}. Watch for "
                       "over-optimism — don't double-count the read *and* the event.")
        with st.expander("💡 Feedback", expanded=True):
            st.markdown(verdict + f" Strong method: anchor on the {base:.0f} base, apply **one** "
                        f"dominant adjustment, and sanity-check against your busiest real day "
                        f"({df['demand'].max()}).")
    reflect("r2_when",
            "Give one situation where qualitative forecasting is the RIGHT choice, and one weakness "
            "of relying on manager opinion.", "When qualitative fits / its weakness",
            "Right choice: no history (new location, brand-new item, first-ever event). Weakness: bias "
            "— anchoring, optimism, or politics distort the number and it's hard to audit.")
    completion(["r2_fc", "r2_when"], "Tab 3 · Naïve")

# ---- 3 Naïve ----
with tabs[3]:
    st.session_state["section"] = "Module 3 — Naïve forecast"
    st.subheader("Module 3 — Naïve forecast")
    objective_box(6, "Compute a naïve forecast and treat it as the benchmark to beat.")
    st.latex(r"F_{t+1} = A_t \quad\text{(next period's forecast = this period's actual)}")
    st.markdown("**Worked data — this week's demand (read it like a spreadsheet):**")
    excel_grid(["A", "B", "C", "D", "E", "F", "G"],
               [["", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
                ["Demand", W["Mon"], W["Tue"], W["Wed"], W["Thu"], W["Fri"], L["sat_actual"]]],
               start_row=4, label_cols=[0])
    naive_cells = {"B5": W["Mon"], "C5": W["Tue"], "D5": W["Wed"], "E5": W["Thu"],
                   "F5": W["Fri"], "G5": L["sat_actual"]}
    st.markdown("**Apply it:** using Friday's actual, what is the naïve forecast for **SATURDAY**?")
    num_task("r3_sat", "Naïve forecast for Saturday", W["Fri"],
             worked_md=f"Next = today's actual. Friday (F5) = {W['Fri']} → Saturday forecast = "
                       f"**{W['Fri']}**.",
             feedback_md=(f"**{W['Fri']}** is the only naïve answer — but is it *good*? Saturdays here "
                          "run ~55% of a weekday, so naïve badly over-forecasts weekends."),
             excel_model="=F5", excel_hint="reference Friday's cell (F5), don't type its number",
             cells=naive_cells),
    st.markdown("**🔁 Iterate:** the naïve forecast **for Thursday** was Wednesday's actual (D5); "
                "Thursday's actual is E5. What is the **absolute error on Thursday**?")
    num_task("r3_err", "Naïve absolute error (Thursday)", abs(W["Thu"] - W["Wed"]),
             worked_md=f"|Actual − Forecast| = |E5 − D5| = |{W['Thu']} − {W['Wed']}| = "
                       f"**{abs(W['Thu']-W['Wed'])}**.",
             feedback_md="Small errors like this are why naïve survives on calm stretches; its trouble "
                         "is *predictable* days (weekends, exams) where error explodes.",
             excel_model="=ABS(E5-D5)", excel_hint="wrap the difference of the two cells in ABS()",
             cells=naive_cells),
    reflect("r3_use", "Why do analysts keep the naïve forecast around even when they have better "
            "models?", "Why keep the naïve benchmark",
            "It's the free baseline: if a complex model can't beat naïve out-of-sample, the complexity "
            "isn't earning its keep. Naïve also updates instantly and needs no history.")
    completion(["r3_sat", "r3_err", "r3_use"], "Tab 4 · Moving Avg")

# ---- 4 Moving average ----
with tabs[4]:
    st.session_state["section"] = "Module 4 — Moving average"
    st.subheader("Module 4 — Moving average")
    objective_box(8, "Compute an n-period moving average; explore the smoothing/lag trade-off.")
    st.latex(r"F_{t+1} = \frac{A_t + A_{t-1} + \dots + A_{t-n+1}}{n}")
    st.markdown("**Worked data:**")
    excel_grid(["A", "B", "C", "D", "E", "F", "G"],
               [["", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
                ["Demand", W["Mon"], W["Tue"], W["Wed"], W["Thu"], W["Fri"], L["sat_actual"]]],
               start_row=4, label_cols=[0])
    ma_cells = {"B5": W["Mon"], "C5": W["Tue"], "D5": W["Wed"], "E5": W["Thu"],
                "F5": W["Fri"], "G5": L["sat_actual"]}
    ma3 = (W["Wed"] + W["Thu"] + W["Fri"]) / 3
    st.markdown("**Apply it:** the **3-day moving-average forecast for SATURDAY**, averaging the last "
                "three actual days (Wed, Thu, Fri = D5, E5, F5).")
    num_task("r4_ma3", "3-day MA forecast for Saturday", round(ma3, 2),
             worked_md=f"$(D5+E5+F5)/3 = ({W['Wed']}+{W['Thu']}+{W['Fri']})/3 = **{ma3:.2f}**$.",
             feedback_md=f"**{ma3:.1f}** is smoother than the naïve {W['Fri']} — it averages three "
                         "days. It still ignores that Saturday is structurally slow, and it **lags** "
                         "real jumps by about n/2 days.",
             excel_model="=AVERAGE(D5:F5)", excel_hint="AVERAGE over the three cells D5:F5",
             cells=ma_cells),
    ma3b = (W["Thu"] + W["Fri"] + L["sat_actual"]) / 3
    st.markdown(f"**🔁 Iterate (roll the window):** Saturday's actual came in at **{L['sat_actual']}** "
                "(G5). Compute the **3-day MA forecast for SUNDAY** using Thu, Fri, Sat (E5, F5, G5).")
    num_task("r4_ma3b", "3-day MA forecast for Sunday", round(ma3b, 2),
             worked_md=f"Drop Wed, add Sat: $(E5+F5+G5)/3 = ({W['Thu']}+{W['Fri']}+{L['sat_actual']})/3"
                       f" = **{ma3b:.2f}**$.",
             feedback_md="See how the window **rolls** — oldest day drops off, newest is added.",
             excel_model="=AVERAGE(E5:G5)", excel_hint="shift the range one cell right (E5:G5)",
             cells=ma_cells),
    st.markdown("### 🔎 Guided exploration — window width")
    st.markdown("**Do this:** set the slider to **2**, then to **8**, and watch the orange line vs. the "
                "blue demand line, and the **MAD** metric.")
    w = st.slider("Moving-average window (days)", 2, 10, 3, key="r4w")
    d = df.copy(); d[f"MA{w}"] = moving_average(d["demand"], w)
    st.line_chart(d.set_index("day")[["demand", f"MA{w}"]])
    mad_w = mad(d["demand"], d[f"MA{w}"])
    best_w = min(range(2, 11), key=lambda k: mad(d["demand"], moving_average(d["demand"], k)))
    best_mad = mad(d["demand"], moving_average(d["demand"], best_w))
    cA, cB = st.columns(2)
    cA.metric(f"MAD at window = {w}", f"{mad_w:.1f}")
    cB.metric("Lowest-MAD window", f"{best_w}", help=f"MAD {best_mad:.1f}")
    if w == 3:
        st.caption("👆 Move the window slider above to a new value — the interpretation and the "
                   "lowest-MAD comparison will appear here once you do.")
    else:
        if w < best_w:
            obs = (f"**What you should see:** a narrow window ({w}) tracks spikes but is jumpy; MAD "
                   f"({mad_w:.0f}) is above the best window's ({best_mad:.0f}).")
        elif w > best_w:
            obs = (f"**What you should see:** a wide window ({w}) is smooth but lags turns and misses "
                   f"peaks — MAD ({mad_w:.0f}) is above the best ({best_mad:.0f}).")
        else:
            obs = f"**Nice — you found the lowest-MAD window ({best_w})** for your data. 🎯"
        st.success(obs + " Rule: wider = smoother but slower; narrower = quicker but noisier.")
        save("r4_explore", "MA exploration", f"tried window {w} (best {best_w})")
    reflect("r4_tradeoff", "Demand jumps for finals week. Would a 2-day or a 4-day moving average catch "
            "the jump faster, and why?", "Window width vs responsiveness",
            "A **2-day** MA reacts faster — recent days carry more weight, so it climbs sooner. The "
            "4-day MA lags because it keeps averaging in old, lower days.")
    completion(["r4_ma3", "r4_ma3b", "r4_explore", "r4_tradeoff"], "Tab 5 · Exp. Smoothing")

# ---- 5 Exponential smoothing ----
with tabs[5]:
    st.session_state["section"] = "Module 5 — Exponential smoothing"
    st.subheader("Module 5 — Exponential smoothing")
    objective_box(8, "Compute smoothing updates (twice, feeding forward) and tune α.")
    st.latex(r"F_{t+1} = F_t + \alpha\,(A_t - F_t)")
    al = L["es_alpha"]; F_t = L["es_F"]; A_t = L["es_A"]; A_t2 = L["es_A2"]
    es1 = F_t + al * (A_t - F_t); es2 = es1 + al * (A_t2 - es1)
    st.markdown("**Worked data — α and two days. The 'New forecast' column is the forecast for the "
                "NEXT day:**")
    excel_grid(["A", "B"], [["alpha", al]], start_row=4, label_cols=[0])
    excel_grid(["A", "B", "C", "D"],
               [["", "Prior F", "Actual", "New F (next day)"],
                ["Day 1", F_t, A_t, "?"],
                ["Day 2", "= D7", A_t2, "?"]],
               start_row=6, label_cols=[0])
    es_cells = {"B4": al, "B7": F_t, "C7": A_t, "C8": A_t2, "D7": round(es1, 6)}
    st.markdown("**Apply it (step 1):** using Day-1's prior forecast (B7), Day-1's actual (C7) and α "
                "(B4), compute **D7 — the forecast for DAY 2**.")
    num_task("r5_es1", "Exp-smoothing forecast for Day 2", round(es1, 2),
             worked_md=f"$D7 = B7 + B4\\,(C7-B7) = {F_t} + {al}({A_t}-{F_t}) = **{es1:.1f}**$.",
             feedback_md=f"**{es1:.0f}** moved only {int(al*100)}% of the way from {F_t} toward {A_t}. "
                         "α is the fraction of the latest surprise you believe.",
             excel_model="=B7+B4*(C7-B7)",
             excel_hint="reference the alpha cell B4 — do not type 0." + f"{int(al*10)}",
             cells=es_cells),
    st.markdown(f"**🔁 Iterate (step 2):** Day-2's actual is C8 = **{A_t2}**. Your Day-2 forecast "
                f"(≈ {es1:.0f}) is now the prior, stored in **D7**. Compute **D8 — the forecast for "
                "DAY 3**.")
    num_task("r5_es2", "Exp-smoothing forecast for Day 3", round(es2, 2),
             worked_md=f"$D8 = D7 + B4\\,(C8-D7) = {es1:.1f} + {al}({A_t2}-{es1:.1f}) = **{es2:.1f}**$.",
             feedback_md="This is the heart of smoothing: **yesterday's forecast becomes today's "
                         "input.** Each past actual echoes with shrinking weight.",
             excel_model="=D7+B4*(C8-D7)", excel_hint="use your step-1 answer cell D7 as the prior",
             cells=es_cells),
    st.markdown("### 🔎 Guided exploration — the constant α")
    st.markdown("**Do this:** drag α to **0.1**, then **0.9**. Watch how tightly the orange forecast "
                "hugs demand and read the MAD; then hunt for the lowest-MAD α.")
    alpha = st.slider("α (smoothing constant)", 0.05, 0.95, 0.30, 0.05, key="r5a")
    d = df.copy(); d["es"] = exp_smoothing(d["demand"], alpha)
    st.line_chart(d.set_index("day")[["demand", "es"]])
    sweep = pd.DataFrame({"alpha": np.round(np.arange(0.1, 0.91, 0.1), 2)})
    sweep["MAD"] = [mad(d["demand"], exp_smoothing(d["demand"], a)) for a in sweep["alpha"]]
    best_a = sweep.loc[sweep["MAD"].idxmin(), "alpha"]; best_am = sweep["MAD"].min()
    c1, c2 = st.columns(2)
    c1.metric(f"MAD at α = {alpha}", f"{mad(d['demand'], d['es']):.1f}")
    c2.metric("Lowest-MAD α", f"{best_a}", help=f"MAD {best_am:.1f}")
    st.bar_chart(sweep.set_index("alpha")["MAD"])
    if alpha == 0.30:
        st.caption("👆 Drag α above to a new value — the interpretation and MAD comparison will "
                   "appear here once you do.")
    else:
        rel = "above" if alpha > best_a else "below" if alpha < best_a else "right at"
        st.success(f"**What you should see:** low α → smooth, sluggish; high α → jumpy, chases "
                   f"wiggles. Your α is {rel} the MAD-minimizing α ({best_a}). Can you beat MAD "
                   f"{best_am:.0f}?")
        save("r5_explore", "α exploration", f"tried α {alpha} (best {best_a})")
    reflect("r5_alpha", "Your demand is noisy day-to-day but has no real trend. Would you pick a high "
            "or low α, and why?", "Choosing α for noisy data",
            "**Low α.** With no true trend, big swings are just noise, so a low α averages them out.")
    completion(["r5_es1", "r5_es2", "r5_explore", "r5_alpha"], "Tab 6 · Seasonality")

# ---- 6 Seasonality (students compute the averages) ----
with tabs[6]:
    st.session_state["section"] = "Module 6 — Seasonality"
    st.subheader("Module 6 — Seasonality (day-of-week)")
    objective_box(10, "Compute the day average and overall average yourself, form the index, apply it.")
    st.latex(r"\text{Index}=\frac{\text{day average}}{\text{overall (grand) average}}\qquad "
             r"F=(\text{base level})\times(\text{index})")
    st.markdown("**Worked data — 4 weeks of daily demand. You will compute the averages.**")
    grid_rows = [["Week", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]]
    for wk in range(4):
        grid_rows.append([f"Wk {wk+1}"] + L["seas_grid"][wk])
    excel_grid(["A", "B", "C", "D", "E", "F", "G", "H"], grid_rows, start_row=4, label_cols=[0])
    st.caption("Weeks are rows 5–8. Mon=B, Tue=C, Wed=D, Thu=E, Fri=F, Sat=G, Sun=H.")
    seas_cells = {}
    for wk in range(4):
        for j, col in enumerate("BCDEFGH"):
            seas_cells[f"{col}{5+wk}"] = L["seas_grid"][wk][j]
    wed_avg, sat_avg, grand, base = L["wed_avg"], L["sat_avg"], L["grand"], L["base"]
    idx_wed = wed_avg / grand; idx_sat = sat_avg / grand; fc_wed = base * idx_wed
    # answer cells students conceptually fill:
    seas_cells.update({"B10": round(wed_avg, 6), "B11": round(sat_avg, 6), "B12": round(grand, 6),
                       "B13": base, "B14": round(idx_wed, 6), "B15": round(idx_sat, 6)})

    st.markdown("**Step 1 — compute the WEDNESDAY average** (average of the four Wednesday values, "
                "cells D5:D8).")
    num_task("r6_wedavg", "Wednesday average demand", round(wed_avg, 2), tol=0.02,
             worked_md=f"$=AVERAGE(D5:D8) = ({'+'.join(str(L['seas_grid'][w][DAYS.index('Wed')]) for w in range(4))})/4 = **{wed_avg:.1f}**$.",
             feedback_md="This is the typical Wednesday. You'll compare it to the overall average next "
                         "to see how much busier Wednesdays run.",
             excel_model="=AVERAGE(D5:D8)", excel_hint="AVERAGE the four Wednesday cells D5:D8",
             cells=seas_cells),
    st.markdown("**Step 2 — compute the OVERALL (grand) average** of all 28 days (cells B5:H8).")
    num_task("r6_grand", "Overall (grand) average demand", round(grand, 2), tol=0.02,
             worked_md=f"$=AVERAGE(B5:H8)$ over all 28 values $= **{grand:.1f}**$.",
             feedback_md="The grand average is the 'typical day' baseline. Because weekends drag it "
                         "down, it sits below the midweek days.",
             excel_model="=AVERAGE(B5:H8)", excel_hint="AVERAGE the whole block B5:H8",
             cells=seas_cells),
    st.markdown("**Step 3 — form the WEDNESDAY seasonal index** = your Wednesday average (B10) ÷ your "
                "overall average (B12).")
    num_task("r6_idx", "Wednesday seasonal index", round(idx_wed, 3), tol=0.03,
             worked_md=f"Index = B10/B12 = {wed_avg:.1f}/{grand:.1f} = **{idx_wed:.3f}**.",
             feedback_md=f"An index **>1** = busy day. {idx_wed:.2f} means Wednesdays run "
                         f"~{abs(idx_wed-1)*100:.0f}% above a typical day.",
             excel_model="=B10/B12", excel_hint="your Wed-average cell ÷ your grand-average cell "
             "(or =AVERAGE(D5:D8)/AVERAGE(B5:H8))", cells=seas_cells),
    st.markdown(f"**Step 4 — apply it:** with base level = B13 ({base}) and your Wednesday index (B14), "
                "forecast **next WEDNESDAY**.")
    num_task("r6_fc", "Seasonal forecast for next Wednesday", round(fc_wed, 1), tol=0.03,
             worked_md=f"$F = B13\\times B14 = {base}\\times{idx_wed:.3f} = **{fc_wed:.0f}**$.",
             feedback_md="Now the forecast respects the weekly rhythm — unlike naïve/MA it won't "
                         "over-serve weekends.",
             excel_model="=B13*B14", excel_hint="base cell × your Wed-index cell", cells=seas_cells),
    st.markdown("**🔁 Iterate:** compute the **SATURDAY average** (G5:G8), then the **Saturday "
                "index** (Sat average B11 ÷ overall average B12).")
    num_task("r6_satavg", "Saturday average demand", round(sat_avg, 2), tol=0.02,
             worked_md=f"$=AVERAGE(G5:G8) = **{sat_avg:.1f}**$.",
             feedback_md="Much lower than midweek — that's the weekend dip.",
             excel_model="=AVERAGE(G5:G8)", excel_hint="AVERAGE the four Saturday cells G5:G8",
             cells=seas_cells),
    num_task("r6_satidx", "Saturday seasonal index", round(idx_sat, 3), tol=0.03,
             worked_md=f"B11/B12 = {sat_avg:.1f}/{grand:.1f} = **{idx_sat:.3f}** — Saturdays run "
                       f"~{(1-idx_sat)*100:.0f}% below typical.",
             feedback_md=f"Index {idx_sat:.2f} × base {base} = **{base*idx_sat:.0f}** for Saturday — "
                         "far below a naïve weekday guess. That gap is the over-staffing seasonality "
                         "prevents.",
             excel_model="=B11/B12", excel_hint="your Sat-average cell ÷ your grand-average cell",
             cells=seas_cells),
    reflect("r6_why", "Why does ignoring day-of-week seasonality make you OVERSTAFF weekends and "
            "UNDERSTAFF midweek?", "Consequence of ignoring seasonality",
            "Flat methods forecast every day near the weekly *average* — above true weekend demand "
            "(idle staff, spoiled fruit) and below midweek peaks (stockouts, lines).")
    completion(["r6_wedavg", "r6_grand", "r6_idx", "r6_fc", "r6_satavg", "r6_satidx", "r6_why"],
               "Tab 7 · Regression")

# ---- 7 Regression ----
with tabs[7]:
    st.session_state["section"] = "Module 7 — Regression"
    st.subheader("Module 7 — Regression (causal forecasting)")
    objective_box(6, "Use a fitted equation to forecast from drivers; read coefficients.")
    st.markdown("Regression links demand to measurable **causes**. Given fitted equation:")
    st.latex(r"\widehat{Demand}=40+2.5\,(Temp)+45\,(Promo)+1.8\,(Attendance)")
    st.markdown("**Worked data — coefficients and two days of drivers:**")
    excel_grid(["A", "B"],
               [["Intercept", REG["intercept"]], ["Temp coef", REG["temp"]],
                ["Promo coef", REG["promo"]], ["Attend coef", REG["attend"]]],
               start_row=4, label_cols=[0])
    excel_grid(["A", "B", "C", "D"],
               [["Temp", "Promo", "Attendance", "Predicted"],
                [L["r1"][0], L["r1"][1], L["r1"][2], "?"],
                [L["r2"][0], L["r2"][1], L["r2"][2], "?"]],
               start_row=9)
    t1, p1, a1 = L["r1"]; t2, p2, a2 = L["r2"]
    pred = REG["intercept"] + REG["temp"] * t1 + REG["promo"] * p1 + REG["attend"] * a1
    pred2 = REG["intercept"] + REG["temp"] * t2 + REG["promo"] * p2 + REG["attend"] * a2
    reg_cells = {"B4": REG["intercept"], "B5": REG["temp"], "B6": REG["promo"], "B7": REG["attend"],
                 "A10": t1, "B10": p1, "C10": a1, "A11": t2, "B11": p2, "C11": a2}
    st.markdown(f"**Apply it:** predict demand for **DAY 1** — Temp {t1} (A10), Promo {p1} (B10), "
                f"Attendance {a1} (C10).")
    num_task("r7_pred", "Regression prediction (Day 1)", round(pred, 1), tol=0.02,
             worked_md=f"$B4+B5\\cdot A10+B6\\cdot B10+B7\\cdot C10 = 40+2.5({t1})+45({p1})+1.8({a1}) "
                       f"= **{pred:.0f}**$.",
             feedback_md=f"**{pred:.0f}** is right. A promo is worth ~45 juices, each °F ~2.5. "
                         "Regression forecasts **new** conditions, but you must know the drivers ahead.",
             excel_model="=B4+B5*A10+B6*B10+B7*C10",
             excel_hint="reference the coefficient cells B4:B7 — don't type 40, 2.5, 45, 1.8",
             cells=reg_cells),
    st.markdown(f"**🔁 Iterate:** re-forecast **DAY 2** — Temp {t2} (A11), Promo {p2} (B11), "
                f"Attendance {a2} (C11).")
    num_task("r7_pred2", "Regression prediction (Day 2)", round(pred2, 1), tol=0.02,
             worked_md=f"$40+2.5({t2})+45({p2})+1.8({a2}) = **{pred2:.1f}**$.",
             feedback_md="Same equation, very different answer — a causal model extrapolates to "
                         "*conditions*, not just past demand.",
             excel_model="=B4+B5*A11+B6*B11+B7*C11", excel_hint="same coefficient cells, row-11 drivers",
             cells=reg_cells),
    reflect("r7_driver", "Which driver would you most want to know accurately the night before, and "
            "why?", "Most valuable driver to know early",
            "Usually **promotion** (you control it, worth ~45 juices) or **temperature** (a reliable "
            "next-day forecast exists).")
    completion(["r7_pred", "r7_pred2", "r7_driver"], "Tab 8 · Accuracy")

# ---- 8 Accuracy (now before Model Selection) ----
with tabs[8]:
    st.session_state["section"] = "Module 8 — Accuracy (MAD & MAPE)"
    st.subheader("Module 8 — Accuracy: MAD & MAPE")
    objective_box(7, "Compute MAD and MAPE by hand and interpret each — the tools you'll use to "
                  "choose a method in Module 9.")
    st.latex(r"MAD=\frac{\sum|A_t-F_t|}{n}\qquad MAPE=\frac{100}{n}\sum\left|\frac{A_t-F_t}{A_t}\right|")
    A = L["acts"]; F = L["fcs"]; errs = [abs(a - f) for a, f in zip(A, F)]
    st.markdown("**Worked data (Actual = column B, Forecast = column C):**")
    excel_grid(["A", "B", "C", "D", "E"],
               [["Day", "Actual", "Forecast", "|Error|", "%Error"]]
               + [[i + 1, A[i], F[i], "?", "?"] for i in range(4)],
               start_row=4)
    acc_cells = {}
    for i in range(4):
        r = 5 + i
        acc_cells[f"B{r}"] = A[i]; acc_cells[f"C{r}"] = F[i]
        acc_cells[f"D{r}"] = abs(A[i] - F[i]); acc_cells[f"E{r}"] = abs(A[i] - F[i]) / A[i]
    mad_ans = float(np.mean(errs))
    mape_ans = float(np.mean([abs(a - f) / a for a, f in zip(A, F)]) * 100)
    st.caption("Build column D as D5 =ABS(B5-C5) down the rows, and column E as E5 =D5/B5 (the "
               "checker knows those values), or compute straight from B and C.")
    st.markdown("**Apply it (MAD):** the Mean Absolute Deviation across all four days.")
    num_task("r9_mad", "MAD across the four days", round(mad_ans, 2), tol=0.03, units="customers",
             worked_md=f"|Errors| = {', '.join(map(str, errs))} → sum {sum(errs)} → "
                       f"MAD = {sum(errs)}/4 = **{mad_ans:.1f}** customers.",
             feedback_md=f"**{mad_ans:.1f} customers** is your average miss in *units* — use it to size "
                         "a safety buffer. A stronger analysis also reports **bias** (average *signed* "
                         "error) to catch consistent over/under-forecasting.",
             excel_model="=AVERAGE(D5:D8)", excel_hint="average the |Error| column D5:D8 "
             "(or =SUMPRODUCT(ABS(B5:B8-C5:C8))/4)", cells=acc_cells),
    st.markdown("**🔁 Iterate (MAPE):** the Mean Absolute Percent Error across all four days.")
    num_task("r9_mape", "MAPE across the four days", round(mape_ans, 2), tol=0.06, units="%",
             worked_md="%errors = " + ", ".join(f"{abs(a-f)/a*100:.1f}%" for a, f in zip(A, F))
                       + f" → average = **{mape_ans:.1f}%**.",
             feedback_md="**~{:.0f}%** is *unit-free*, so you can compare a slow Saturday with a busy "
                         "exam day. Report MAD **and** MAPE together.".format(mape_ans),
             excel_model="=AVERAGE(E5:E8)*100", excel_hint="average the %Error column E5:E8, ×100",
             cells=acc_cells),
    reflect("r9_interpret", "In one sentence each: what does MAD tell you that MAPE doesn't, and what "
            "does MAPE tell you that MAD doesn't?", "Interpreting MAD vs MAPE",
            "MAD is in **customers**, so it directly sizes buffers/staffing. MAPE is a **percent**, so "
            "it compares accuracy fairly across busy and slow days regardless of scale.")
    completion(["r9_mad", "r9_mape", "r9_interpret"], "Tab 9 · Model Selection")

# ---- 9 Model selection (student identifies lowest error, then selects) ----
with tabs[9]:
    st.session_state["section"] = "Module 9 — Model selection"
    st.subheader("Module 9 — Choosing a method")
    objective_box(6, "Read the hold-out errors, identify the lowest, and select a method to use.")
    st.markdown("Each method was fit on the early term and scored on the **last 14 days it never "
                "saw**. Lower MAD = better out-of-sample accuracy. **Read the table and find the "
                "winner yourself.**")
    d = df.copy()
    d["naïve"] = naive_forecast(d["demand"]); d["MA3"] = moving_average(d["demand"], 3)
    d["exp. smoothing"] = exp_smoothing(d["demand"], 0.3); d["seasonal"] = seasonal_naive_forecast(d)
    beta, _ = regression_fit(train)
    Xte = np.column_stack([np.ones(len(df)), df["temp_f"], df["promo"], df["attendance"],
                           df["exam"], df["event"]])
    d["regression"] = Xte @ beta
    te = d.iloc[-holdout:]
    methods = ["naïve", "MA3", "exp. smoothing", "seasonal", "regression"]
    score = pd.DataFrame([{"method": m, "MAD": round(mad(te["demand"], te[m]), 1),
                           "MAPE %": round(mape(te["demand"], te[m]), 1),
                           "RMSE": round(rmse(te["demand"], te[m]), 1)} for m in methods])
    # Display in method order (NOT sorted) so students must find the minimum.
    st.dataframe(score, use_container_width=True, hide_index=True)
    st.bar_chart(score.set_index("method")["MAD"])
    winner = score.sort_values("MAD").iloc[0]["method"]

    st.markdown("**Step 1 — identify the winner.** Which method has the **lowest MAD**?")
    guess = st.selectbox("Method with the lowest MAD", ["— choose —"] + methods, key="r9id",
                         label_visibility="collapsed")
    if guess != "— choose —":
        ok = (guess == winner)
        best_mad = score.sort_values("MAD").iloc[0]["MAD"]
        (st.success if ok else st.error)(
            f"{'✅ Correct' if ok else '❌ Look again'} — the lowest MAD is **{best_mad}**, which is "
            f"**{winner}**." + ("" if ok else f" You chose {guess}."))
        record_first("msel_identify", ok)
        save("msel_identify", "Identified lowest-MAD method", f"{guess}", ok)

    st.markdown("**Step 2 — select the method** you'll carry into the decision round (usually the "
                "lowest-error one, but you may justify another).")
    pick = st.selectbox("Method for next period", ["— choose —"] + methods, key="r9pick",
                        label_visibility="collapsed")
    if pick != "— choose —":
        st.session_state["chosen_method"] = pick
        save("msel_pick", "Method selected for next period", pick)
        st.info(f"You'll plan against **{pick}** when you run {BAR_NAME}.")

    reflect("msel_defend", "Defend your selection using its MAD and MAPE, and say why it beats the "
            "naïve benchmark (2–3 sentences).", "Defense of chosen method",
            "A strong defense cites the *hold-out* MAD/MAPE, compares to naïve, and connects the "
            "numbers to the decision (smaller MAD → smaller buffer, less waste and fewer stockouts). "
            "Seasonal/regression win here because demand is driven by knowable factors.")
    completion(["msel_identify", "msel_pick", "msel_defend"], f"Tab 🏪 Run {BAR_NAME}")

# ---- Run the Bar (student calculates the plan, then implements) ----
with tabs[10]:
    st.session_state["section"] = f"Run {BAR_NAME}"
    st.subheader(f"🏪 Run {BAR_NAME} — forecast, calculate the plan, then implement")
    objective_box(6, "Choose a forecast, CALCULATE each plan number yourself, then run the day.")
    method = st.session_state.get("chosen_method", st.session_state.get("best_method", "seasonal"))
    d = df.copy()
    if method == "seasonal":
        suggested = seasonal_naive_forecast(d).iloc[-1]
    elif method == "exp. smoothing":
        suggested = exp_smoothing(d["demand"], 0.3).iloc[-1]
    elif method == "MA3":
        suggested = moving_average(d["demand"], 3).iloc[-1]
    elif method == "naïve":
        suggested = d["demand"].iloc[-1]
    else:
        beta, _ = regression_fit(train); Lr = df.iloc[-1]
        suggested = (beta[0] + beta[1]*Lr["temp_f"] + beta[2]*Lr["promo"] + beta[3]*Lr["attendance"]
                     + beta[4]*Lr["exam"] + beta[5]*Lr["event"])
    fc0 = int(max(0, suggested))
    promo_be = int(PROMO_COST / (CONTRIB * PROMO_LIFT))

    st.markdown("#### Step 1 — Choose your forecast")
    st.markdown(f"For reference, your method (**{method}**) suggests ≈ **{fc0}**. **You** decide the "
                "number to plan against for **tomorrow**.")
    cA, cB = st.columns([2, 1])
    fc_choice = cA.number_input("➡️ Your demand forecast for tomorrow (customers)", 0, 1000,
                                value=int(fc0), key="rb_fc")
    if cB.button("Lock in forecast ▶", type="primary"):
        st.session_state["plan_fc"] = int(fc_choice)
        autosave()

    if "plan_fc" in st.session_state:
        f = st.session_state["plan_fc"]
        rec_emp = int(np.ceil(f / EMP_DAILY_CAP)); rec_bot = int(f * 0.30)
        rec_fruit = f - rec_bot; rec_mob = int(f * 0.25)
        st.markdown(f"#### Step 2 — Calculate each plan number from your forecast of **{f}**")
        st.caption("Work out each value with the formula shown, then check it. 1 employee serves "
                   f"{EMP_DAILY_CAP} customers/day.")
        plan_calc("emp", "Employees to schedule", rec_emp,
                  f"round UP({f} ÷ {EMP_DAILY_CAP})")
        plan_calc("bot", "Bottles to prepackage", rec_bot, f"0.30 × {f}")
        plan_calc("fruit", "Fresh fruit servings", rec_fruit, f"{f} − bottles")
        plan_calc("mob", "Mobile capacity to reserve", rec_mob, f"0.25 × {f}")
        st.markdown(f"**Promotion?** A promo lifts demand ~{int(PROMO_LIFT*100)}%; extra margin "
                    f"{PROMO_LIFT:g}×{f}×${CONTRIB:.2f} beats the ${PROMO_COST:.0f} cost when your "
                    f"forecast exceeds **{promo_be}**. Your forecast is {f} → "
                    f"**{'run the promo' if f > promo_be else 'skip the promo'}**.")

        st.markdown("#### Step 3 — Implement your plan and open")
        st.caption("Set the sliders (start from your calculated numbers), then open the bar.")
        c1, c2, c3 = st.columns(3)
        employees = c1.slider("Employees scheduled", 1, 12, value=rec_emp, key="rb_emp")
        bottles = c2.slider("Bottles prepackaged", 0, 500, value=rec_bot, key="rb_bot")
        fruit_prep = c3.slider("Fruit servings prepped", 0, 1000, value=rec_fruit, key="rb_fruit")
        c4, c5 = st.columns(2)
        mobile = c4.slider("Mobile-order capacity reserved", 0, 400, value=rec_mob, key="rb_mob")
        promo = c5.checkbox("Run a promotion", value=f > promo_be, key="rb_promo")
        rng = np.random.default_rng(seed + 99)
        actual = int(max(0, rng.normal(fc0, fc0 * 0.09)))
        if st.button(f"▶ Open {BAR_NAME} for the day"):
            r = run_day_pnl(actual, employees, fruit_prep, bottles, promo, mobile)
            # Performance vs. an ideal plan that perfectly matches the demand that showed up.
            dem = r["realized_demand"]; ib = int(dem * 0.30)
            ideal = run_day_pnl(dem, max(1, int(np.ceil(dem / EMP_DAILY_CAP))),
                                dem - ib, ib, False, int(dem * 0.25))["profit"]
            prof_ratio = 100 * r["profit"] / ideal if ideal > 0 else 0
            perf_run = round(0.6 * max(0, min(100, prof_ratio)) + 0.4 * r["satisfaction"], 1)
            st.session_state.setdefault("runs", []).append(
                {"forecast": f, "actual": r["realized_demand"], "error": f - r["realized_demand"],
                 "profit": r["profit"], "satisfaction": r["satisfaction"], "perf": perf_run,
                 "ideal": round(ideal, 2), "gap": round(max(0.0, ideal - r["profit"]), 2)})
            a, b, c, e = st.columns(4)
            a.metric("Actual demand", r["realized_demand"]); b.metric("Units sold", r["units_sold"])
            c.metric("Unmet demand", r["unmet_demand"]); e.metric("Satisfaction", f"{r['satisfaction']}%")
            pcol1, pcol2 = st.columns(2)
            pcol1.metric("💵 Profit for the day", f"${r['profit']:.2f}")
            pcol2.metric("⭐ Day performance", f"{perf_run}/100",
                         help="60% how close your profit is to a perfect-match plan, 40% satisfaction.")
            st.bar_chart(pd.DataFrame({"amount": [
                r["revenue"], -r["labor_cost"], -r["ingredient_cost"], -r["fruit_waste"],
                -r["bottle_waste"], -r["promo_cost"], -r["goodwill_lost"]]},
                index=["Revenue", "Labor", "Ingredients", "Fruit waste", "Bottle waste", "Promo",
                       "Lost goodwill"]))
            ideal_emp = int(np.ceil(r["realized_demand"] / EMP_DAILY_CAP)); tips = []
            if r["fruit_waste"] + r["bottle_waste"] > 30:
                tips.append(f"Wasted ${r['fruit_waste']+r['bottle_waste']:.0f} of prep — demand was "
                            f"{r['realized_demand']}, so ~{r['realized_demand']} servings was enough.")
            if r["goodwill_lost"] > 0:
                tips.append(f"Lost {r['unmet_demand']} sales. Ideal for {r['realized_demand']} "
                            f"customers: {ideal_emp} staff and ≥{r['realized_demand']} servings.")
            if employees > ideal_emp + 1:
                tips.append(f"Scheduled {employees} staff but {ideal_emp} could serve demand — "
                            f"~${(employees-ideal_emp)*SHIFT_HOURS*WAGE_PER_HR:.0f} idle labor.")
            err = f - r["realized_demand"]
            head = ("Tight forecast — low waste *and* low lost sales. 🎯" if abs(err) <= 40 else
                    (f"You **over-forecast by {err}**." if err > 0 else
                     f"You **under-forecast by {-err}**."))
            (st.success if abs(err) <= 40 else st.warning)(head)
            for t in tips:
                st.markdown(f"- {t}")
            st.markdown("**Reasonable vs. better:** a lost sale costs "
                        f"${CONTRIB+SATISFACTION_PENALTY:.2f} (margin + goodwill) while an extra fruit "
                        f"serving wastes only ${FRUIT_PREP_COST:.2f}, so leaning slightly over on "
                        "fruit is the cheaper mistake.")
            save("rb_result", "Run-the-Bar result",
                 f"forecast {f}, actual {r['realized_demand']}, profit ${r['profit']}, "
                 f"satisfaction {r['satisfaction']}%")
    else:
        st.info("Enter a forecast above and click **Lock in forecast ▶** to start calculating your plan.")

    reflect("rb_lesson", "Based on your P&L, is it more expensive to over- or under-forecast here, and "
            "what does that imply about how you'd bias your plan?", "Cost of over- vs under-forecasting",
            f"Under-forecasting costs the ${CONTRIB:.2f} lost margin **plus** ${SATISFACTION_PENALTY:.2f} "
            f"goodwill per unmet customer; over-prepping fruit wastes only ${FRUIT_PREP_COST:.2f}. So a "
            "small over-prep is the cheaper mistake — set a **service buffer** above the point forecast, "
            "sized by the MAD from Module 8.")
    completion(["rb_result", "rb_lesson"], "Tab 🎓 Debrief")

# ---- Debrief (What? / So what? / Now what?) ----
with tabs[11]:
    st.session_state["section"] = "Debrief"
    st.subheader("🎓 Debrief — What? · So what? · Now what?")
    st.markdown("This is where the experience turns into a lesson you keep. Spend five minutes here — "
                "it's the highest-value part of the whole lab.")
    runs = st.session_state.get("runs", [])
    method_sel = st.session_state.get("chosen_method", "(not selected)")
    done_steps, total_steps = overall_progress()
    comp = int(round(100 * done_steps / total_steps))
    mscore, mcorrect, mtot = mastery_score()
    pscore, _runs = performance_score()

    st.markdown("### 1) What happened?")
    if runs:
        worst = max(runs, key=lambda rr: abs(rr["error"]))
        direction = ("over-forecast" if worst["error"] > 0
                     else "under-forecast" if worst["error"] < 0 else "matched demand")
        st.markdown(f"- You selected the **{method_sel}** method.")
        st.markdown(f"- **Biggest forecast miss:** you planned for **{worst['forecast']}** customers "
                    f"but **{worst['actual']}** showed up — you **{direction}** by "
                    f"**{abs(worst['error'])}**.")
        st.markdown(f"- **Dollar cost of that miss:** you earned **${worst['profit']:.0f}** that day "
                    f"vs. **${worst['ideal']:.0f}** for a perfectly-matched plan — about "
                    f"**${worst['gap']:.0f}** left on the table.")
        bias = float(np.mean([r["error"] for r in runs]))
        tend = ("tended to OVER-forecast" if bias > 5 else
                "tended to UNDER-forecast" if bias < -5 else "were well-balanced")
        st.markdown(f"- Across **{len(runs)} day(s)** you **{tend}** (average miss {bias:+.0f} "
                    "customers).")
    else:
        st.info("Run Juicetification at least once (Tab 🏪) and your biggest miss and its dollar cost "
                "will appear here automatically.")
    st.caption(f"Your scores — Completeness {comp}/100 · "
               f"Mastery {mscore if mscore is not None else '—'}/100 · "
               f"Performance {pscore if pscore is not None else '—'}/100")

    st.markdown("### 2) So what?")
    st.markdown(f"Remember the asymmetry: **under**-forecasting costs a lost sale "
                f"(~${CONTRIB + SATISFACTION_PENALTY:.2f} in margin + goodwill), while "
                f"**over**-forecasting a fresh serving wastes only ~${FRUIT_PREP_COST:.2f}. The two "
                "mistakes are not equally expensive.")
    reflect("dbf_sowhat",
            "Why did your biggest miss happen, and what did it cost the business? Tie it to over- vs. "
            "under-forecasting and to the method you chose.", "Debrief — So what?",
            "A strong answer names the *cause* of the miss (e.g., missed a weekend seasonal dip, or an "
            "event you didn't price in), states the *dollar consequence*, and notes whether the "
            "cheaper or the more expensive kind of error occurred.",
            rubric=["Names why the miss happened", "States the dollar / service cost",
                    "Says whether it was the cheap or the expensive kind of error"])

    st.markdown("### 3) Now what?")
    st.markdown("Turn the experience into a portable rule. Examples: *“Add a buffer of about one MAD "
                "above the point forecast during exam weeks,”* *“Never staff to the average on "
                "weekends — apply the seasonal index,”* or *“Lean slightly toward over-prepping cheap "
                "fruit rather than risking stockouts.”*")
    reflect("dbf_nowhat",
            "Write ONE forecasting rule you'll carry into a real operation — and one sentence on why "
            "it works.", "Debrief — One rule I'll keep",
            "A strong rule is specific and actionable (a number, a trigger, or a clear policy), and "
            "its juicetification ties back to accuracy (MAD/bias) or the cost asymmetry — not just "
            "“forecast better.”",
            rubric=["The rule is specific/actionable (a number, trigger, or policy)",
                    "The reason ties to accuracy or the cost asymmetry",
                    "It would transfer to a real operation, not just this game"])
    if "dbf_nowhat" in st.session_state["responses"]:
        st.success("That rule is the single most valuable thing to walk away with. Nice work, manager.")
    completion(["dbf_sowhat", "dbf_nowhat"], "Tab 📝 Final Report")

# ---- Final Report ----
with tabs[12]:
    st.session_state["section"] = "Report"
    st.subheader("📝 Final Report — review, then submit for grading")
    st.markdown("You've reached the end. **Nothing is locked in** — you can revisit any tab above to "
                "review or change an answer, and it updates here automatically. When you're satisfied, "
                "generate your PDF and upload it to your LMS.")
    reflect("cap_expand", "Capstone (long-range): the college may let you open a SECOND juice bar "
            "across campus. Using demand drivers and forecast accuracy, what evidence would you gather "
            "before recommending it? (4–5 sentences)", "Capstone — second location",
            "A strong answer treats it as a long-range forecast: estimate the new site's demand from "
            "its drivers (foot traffic/attendance nearby, events), check whether current peak demand is "
            "capacity-constrained (unmet demand = room to grow), and quantify uncertainty with MAD/MAPE "
            "before committing capital.", height=140)

    # ---- Three scores: effort, skill, application ----
    resp = st.session_state["responses"]
    done_steps, total_steps = overall_progress()
    comp = int(round(100 * done_steps / total_steps))
    mscore, mcorrect, mtot = mastery_score()
    pscore, runs = performance_score()
    st.markdown("#### Your scores")
    s1, s2, s3 = st.columns(3)
    s1.metric("Completeness", f"{comp} / 100",
              help="How much of the lab you finished — this is effort.")
    s2.metric("Mastery", f"{mscore} / 100" if mscore is not None else "—",
              help="Share of calculations & Excel formulas correct on the FIRST try — this is skill.")
    s3.metric("Performance", f"{pscore} / 100" if pscore is not None else "—",
              help="How well your plan matched demand while running Juicetification — this is application.")
    st.progress(done_steps / total_steps, text=f"Completeness: {done_steps} of {total_steps} steps")
    notes = []
    if mscore is not None:
        notes.append(f"Mastery from {mcorrect}/{mtot} first-try correct.")
    else:
        notes.append("Mastery appears once you check a calculation or formula.")
    if pscore is not None:
        notes.append(f"Performance from {len(runs)} day(s): cumulative profit "
                     f"${sum(r['profit'] for r in runs):.0f}, avg satisfaction "
                     f"{np.mean([r['satisfaction'] for r in runs]):.0f}%.")
    else:
        notes.append("Run Juicetification at least once to earn a Performance score.")
    st.caption("  ·  ".join(notes))
    st.info("**Completeness is effort, Mastery is skill, Performance is applying it well** — they're "
            "scored separately on purpose, so finishing everything doesn't hide a shaky calculation.")

    st.markdown("#### ✅ Review checklist")
    st.caption("Anything marked *needs review* can be finished by clicking that tab above.")
    done_mods = 0
    rows = []
    for name_m, reqs in MODULES:
        have = sum(1 for q in reqs if q in resp)
        complete = have == len(reqs)
        done_mods += complete
        rows.append({"Section": name_m, "Done": f"{have}/{len(reqs)}",
                     "Status": "✅ complete" if complete else "⬜ needs review"})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    if done_mods == len(MODULES):
        st.success("🎉 All sections complete — Completeness 100/100. You're ready to submit.")
    else:
        st.warning(f"{done_mods}/{len(MODULES)} sections complete. Finish the rest for Completeness 100.")

    # Record completion once, when the lab is finished (no-op unless storage is enabled).
    if store.enabled() and sid and done_mods == len(MODULES) and not st.session_state.get("_completion_recorded"):
        import hashlib
        _code = hashlib.sha256(
            f"{CFG['completion_salt']}|{game}|{sid}".encode()).hexdigest()[:8].upper()
        try:
            store.record_completion(game, sid, completion_code=_code, score=comp,
                                    extra={"mastery": mscore, "performance": pscore})
        except Exception:
            pass
        st.session_state["_completion_recorded"] = True
        st.session_state["_completion_code"] = _code
    if store.enabled() and sid and st.session_state.get("_completion_code"):
        st.caption(f"Completion code: {st.session_state['_completion_code']}  ·  "
                   "recorded for your instructor")

    # ---- Submission PDF ----
    name = st.session_state["student"].strip() or "(unnamed)"
    method_sel = st.session_state.get("chosen_method", "(not selected)")
    st.markdown("#### 📄 Submit to your LMS")
    if not st.session_state["student"].strip():
        st.info("Tip: add your name/ID in the sidebar so it appears on the PDF.")
    mastery_line = (f"Mastery score: {mscore}/100 ({mcorrect}/{mtot} first-try correct)"
                    if mscore is not None else "Mastery score: not attempted")
    perf_line = (f"Performance score: {pscore}/100 ({len(runs)} day(s), cumulative profit "
                 f"${sum(r['profit'] for r in runs):.0f})" if pscore is not None
                 else "Performance score: not attempted")
    try:
        pdf_bytes = build_pdf(name, seed, method_sel, st.session_state["order"], resp,
                              mastery_line, perf_line)
        st.download_button("⬇️ Download submission PDF", pdf_bytes,
                           file_name=f"Juicetification_ForecastFrenzy_{re.sub(r'[^A-Za-z0-9]+','_',name)}_{seed}.pdf",
                           mime="application/pdf", type="primary")
    except Exception as ex:  # pragma: no cover
        st.error(f"Couldn't build the PDF ({ex}). The Markdown/CSV below still work for submission.")

    with st.expander("Preview what's in your PDF"):
        checks = [v for v in resp.values() if v["correct"] is not None]
        n_ok = sum(1 for v in checks if v["correct"])
        lines = ["# Juicetification: Forecast Frenzy — Final Report", f"**Student:** {name}  ",
                 f"**Session ID:** {seed}  ", f"**Date:** {dt.date.today().isoformat()}  ",
                 f"**Method selected:** {method_sel}  ",
                 (f"**Checked calculations:** {n_ok}/{len(checks)} correct" if checks
                  else "**Checked calculations:** none attempted"), ""]
        cur = None
        for qid in st.session_state["order"]:
            v = resp[qid]
            if v["section"] != cur:
                cur = v["section"]; lines.append(f"\n## {cur}")
            mark = "" if v["correct"] is None else (" ✅" if v["correct"] else " ❌")
            lines.append(f"- **{v['label']}:** {v['answer']}{mark}")
        st.markdown("\n".join(lines))

    with st.expander("Other formats (Markdown / CSV)"):
        report_md = "\n".join(lines)
        st.download_button("⬇️ Markdown", report_md,
                           file_name=f"juicetification_forecast_frenzy_report_{seed}.md",
                           mime="text/markdown")
        st.download_button("⬇️ Responses CSV",
                           pd.DataFrame([{"section": resp[q]['section'], "item": resp[q]['label'],
                                          "answer": resp[q]['answer'], "correct": resp[q]['correct']}
                                         for q in st.session_state["order"]]).to_csv(index=False),
                           file_name=f"juicetification_forecast_frenzy_responses_{seed}.csv",
                           mime="text/csv")

st.divider()
st.caption("Juicetification: Forecast Frenzy · guided experiential lab · all figures illustrative")
